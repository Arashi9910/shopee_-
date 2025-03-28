#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
紀錄輸出模組 - 負責記錄商品調整資訊並輸出Excel報表

主要功能：
1. 記錄每次調整價格的操作資訊
2. 將記錄輸出為Excel格式的報表
3. 不干擾其他模組的運作
"""

import os
import time
import logging
from datetime import datetime
import pandas as pd
from selenium.webdriver.common.by import By

# 設置日誌
logger = logging.getLogger("紀錄輸出")

class 紀錄管理器:
    """管理調價紀錄並輸出Excel報表"""
    
    def __init__(self):
        """初始化紀錄管理器"""
        # 儲存所有調整記錄的列表
        self.調整紀錄列表 = []
        
        # 設置報表目錄
        self.報表目錄 = "報表"
        self._確保目錄存在(self.報表目錄)
        
        logger.info("紀錄管理器已初始化")
    
    def _確保目錄存在(self, 目錄路徑):
        """確保指定的目錄存在，若不存在則創建"""
        if not os.path.exists(目錄路徑):
            try:
                os.makedirs(目錄路徑)
                logger.info(f"已創建目錄: {目錄路徑}")
            except Exception as e:
                logger.error(f"無法創建目錄 {目錄路徑}: {str(e)}")
    
    def 記錄價格調整(self, 商品名稱, 規格名稱, 原價格, 新價格, 成功=True, 參考規格="", 參考折扣價=None):
        """記錄單次價格調整操作
        
        參數:
            商品名稱 (str): 商品名稱
            規格名稱 (str): 商品規格名稱
            原價格 (str/float): 調整前的原始價格
            新價格 (str/float): 設定的新價格
            成功 (bool): 調整操作是否成功，預設為True
            參考規格 (str): 用於參考價格的規格名稱
            參考折扣價 (str/float): 參考規格的折扣價格
        """
        # 取得當前時間
        當前時間 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 創建記錄
        記錄 = {
            "調整時間": 當前時間,
            "商品名稱": 商品名稱,
            "規格名稱": 規格名稱,
            "原價格": 原價格,
            "參考折扣價": 參考折扣價 if 參考折扣價 else "",
            "設定價格": 新價格,
            "是否成功": "成功" if 成功 else "失敗"
        }
        
        # 檢查原價格和設定價格是否一致
        try:
            原價格值 = float(str(原價格).replace("NT$", "").strip())
            新價格值 = float(str(新價格).replace("NT$", "").strip())
            價格一致 = abs(原價格值 - 新價格值) < 0.01  # 允許微小誤差
            記錄["需檢查"] = "否" if 價格一致 or 原價格 == "未知" or 原價格 == "未設置" else "是"
        except (ValueError, TypeError):
            記錄["需檢查"] = "否"  # 無法轉換為數字時默認不需檢查
        
        # 添加到記錄列表
        self.調整紀錄列表.append(記錄)
        logger.info(f"已記錄價格調整: 商品 '{商品名稱}' 規格 '{規格名稱}' 從 {原價格} 調整至 {新價格}")
        
        return 記錄
    
    def 捕獲價格調整(self, driver, 商品名稱, 規格名稱, 新價格, 成功=True):
        """從頁面捕獲並記錄價格調整
        
        參數:
            driver: Selenium WebDriver實例
            商品名稱 (str): 商品名稱
            規格名稱 (str): 商品規格名稱
            新價格 (str/float): 設定的新價格
            成功 (bool): 調整操作是否成功，預設為True
            
        返回:
            dict: 記錄信息
        """
        # 嘗試從頁面獲取原價格
        原價格 = "未知"
        try:
            # 使用JavaScript查找規格行並獲取當前價格
            原價格 = driver.execute_script("""
                // 找到包含指定規格名稱的行
                const specNameElems = Array.from(document.querySelectorAll('.ellipsis-content.single'));
                const specElem = specNameElems.find(el => el.textContent.trim() === arguments[0]);
                
                if (!specElem) return '未找到';
                
                // 向上查找規格容器
                let specRow = specElem;
                for (let i = 0; i < 5; i++) {
                    if (!specRow) break;
                    if (specRow.classList && 
                       (specRow.classList.contains('discount-edit-item-model-component') || 
                        specRow.classList.contains('discount-view-item-model-component'))) {
                        break;
                    }
                    specRow = specRow.parentElement;
                }
                
                if (!specRow) return '未找到容器';
                
                // 查找價格輸入框或顯示元素
                const priceInput = specRow.querySelector('input.eds-input__input');
                if (priceInput) {
                    return priceInput.value || '無值';
                }
                
                // 如果沒有輸入框，查找顯示的價格元素
                const priceDisplay = specRow.querySelector('.currency-input, [class*="price"]');
                if (priceDisplay) {
                    return priceDisplay.textContent.trim();
                }
                
                return '未找到價格';
            """, 規格名稱)
            
            logger.info(f"從頁面捕獲原價格: {原價格}")
        except Exception as e:
            logger.warning(f"無法從頁面獲取原價格: {str(e)}")
        
        # 記錄調整操作
        return self.記錄價格調整(商品名稱, 規格名稱, 原價格, 新價格, 成功)
    
    def 批量記錄(self, 記錄列表):
        """批量添加多個記錄
        
        參數:
            記錄列表 (list): 包含多個記錄字典的列表
        """
        if not 記錄列表:
            logger.warning("批量記錄: 記錄列表為空")
            return
            
        self.調整紀錄列表.extend(記錄列表)
        logger.info(f"已批量添加 {len(記錄列表)} 條記錄")
    
    def 批量添加記錄(self, 記錄):
        """添加單條或多條記錄
        
        參數:
            記錄: 單個記錄字典或記錄列表
        """
        if isinstance(記錄, list):
            self.批量記錄(記錄)
        elif isinstance(記錄, dict):
            self.調整紀錄列表.append(記錄)
            logger.info(f"已添加單條記錄: {記錄.get('商品名稱', '未知商品')} - {記錄.get('規格名稱', '未知規格')}")
        else:
            logger.warning(f"無法識別的記錄格式: {type(記錄)}")
    
    def 清空記錄(self):
        """清空所有記錄"""
        記錄數量 = len(self.調整紀錄列表)
        self.調整紀錄列表 = []
        logger.info(f"已清空 {記錄數量} 條記錄")
    
    def 輸出Excel報表(self, 檔名前綴="調價紀錄"):
        """將記錄輸出為Excel報表
        
        參數:
            檔名前綴 (str): 報表檔名的前綴，預設為"調價紀錄"
            
        返回:
            str: 報表檔案路徑，若輸出失敗則返回None
        """
        if not self.調整紀錄列表:
            logger.warning("沒有記錄可輸出")
            return None
            
        try:
            # 產生檔名，包含日期時間
            時間戳 = datetime.now().strftime("%Y%m%d_%H%M%S")
            檔名 = f"{檔名前綴}_{時間戳}.xlsx"
            檔案路徑 = os.path.join(self.報表目錄, 檔名)
            
            # 將記錄轉換為DataFrame
            df = pd.DataFrame(self.調整紀錄列表)
            
            # 確保列順序
            列順序 = ["調整時間", "商品名稱", "規格名稱", "原價格", "參考折扣價", "設定價格", "是否成功", "需檢查"]
            有效列 = [col for col in 列順序 if col in df.columns]
            
            # 如果有必要，調整DataFrame列
            if set(df.columns) != set(有效列):
                # 處理列名不匹配的情況
                欄位映射 = {
                    "時間": "調整時間",
                    "商品": "商品名稱",
                    "規格": "規格名稱",
                    "原價": "原價格",
                    "原折扣價": "原價格",
                    "折扣價格": "原價格", 
                    "參考來源": "參考折扣價",
                    "同類規格": "參考折扣價",
                    "參考規格": "參考折扣價",
                    "參考規格折扣價": "參考折扣價",
                    "新價": "設定價格",
                    "新價格": "設定價格",
                    "輸入價格": "設定價格",
                    "結果": "是否成功",
                    "調整結果": "是否成功"
                }
                
                # 重命名列
                for 舊名, 新名 in 欄位映射.items():
                    if 舊名 in df.columns and 新名 not in df.columns:
                        df = df.rename(columns={舊名: 新名})
                
                # 再次檢查有效列
                有效列 = [col for col in 列順序 if col in df.columns]
            
            # 重新排序並選擇需要的列
            if 有效列:
                df = df[有效列]
                
            # 輸出Excel
            with pd.ExcelWriter(檔案路徑, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='調價紀錄', index=False)
                
                # 設置列寬
                worksheet = writer.sheets['調價紀錄']
                列寬對照 = {
                    '調整時間': 20,
                    '商品名稱': 35,
                    '規格名稱': 25,
                    '原價格': 12,
                    '參考折扣價': 15,
                    '設定價格': 12,
                    '是否成功': 10,
                    '需檢查': 10
                }
                
                # 設置各列寬度和格式
                for i, 列名 in enumerate(df.columns):
                    col_letter = chr(65 + i)  # A, B, C, ...
                    if 列名 in 列寬對照:
                        worksheet.column_dimensions[col_letter].width = 列寬對照[列名]
                
                # 指定需檢查的列用不同顏色標記
                if '需檢查' in df.columns:
                    from openpyxl.styles import PatternFill
                    黃色填充 = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    
                    檢查列號 = list(df.columns).index('需檢查') + 1  # Excel列從1開始
                    for row_idx, 需檢查 in enumerate(df['需檢查'], start=2):  # 第2行開始是數據
                        if 需檢查 == '是':
                            # 標記該行的原折扣價和設定價格
                            原折扣價列號 = list(df.columns).index('原折扣價') + 1
                            設定價格列號 = list(df.columns).index('設定價格') + 1
                            需檢查儲存格 = worksheet.cell(row=row_idx, column=檢查列號)
                            
                            需檢查儲存格.fill = 黃色填充
            
            logger.info(f"已輸出Excel報表: {檔案路徑}")
            return 檔案路徑
            
        except Exception as e:
            logger.error(f"輸出Excel報表時發生錯誤: {str(e)}")
            import traceback
            logger.error(f"詳細錯誤: {traceback.format_exc()}")
            return None
    
    def 添加統計摘要(self, 報表檔案路徑, 總處理數=0, 成功數=0, 失敗數=0):
        """在Excel報表中添加統計摘要
        
        參數:
            報表檔案路徑: Excel檔案路徑
            總處理數: 總處理規格數
            成功數: 成功調整的規格數
            失敗數: 失敗的規格數
        """
        try:
            # 讀取現有Excel
            df = pd.read_excel(報表檔案路徑)
            
            # 創建新的Excel工作簿
            with pd.ExcelWriter(報表檔案路徑, engine='openpyxl', mode='a') as writer:
                # 創建摘要表
                摘要資料 = pd.DataFrame([
                    ["調整紀錄摘要", ""],
                    ["總處理規格數", 總處理數],
                    ["成功調整數", 成功數],
                    ["失敗調整數", 失敗數],
                    ["成功率", f"{(成功數/總處理數*100):.1f}%" if 總處理數 > 0 else "0%"]
                ])
                
                # 將摘要表寫入新的工作表
                摘要資料.to_excel(writer, sheet_name="摘要", index=False, header=False)
            
            logger.info(f"已添加統計摘要到報表: {報表檔案路徑}")
            return True
        except Exception as e:
            logger.error(f"添加統計摘要失敗: {str(e)}")
            return False

# 測試代碼
if __name__ == "__main__":
    # 配置日誌
    logging.basicConfig(level=logging.INFO, 
                      format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    
    # 創建紀錄管理器
    紀錄器 = 紀錄管理器()
    
    # 添加一些測試數據
    紀錄器.記錄價格調整("測試商品1", "規格A", "NT$100", "85", True)
    紀錄器.記錄價格調整("測試商品1", "規格B", "NT$120", "99", False)
    紀錄器.記錄價格調整("測試商品2", "規格X", "NT$200", "150", True)
    
    # 輸出Excel
    excel_path = 紀錄器.輸出Excel報表("測試報表")
    
    if excel_path:
        紀錄器.添加統計摘要(excel_path)
        print(f"測試完成，Excel報表已生成: {excel_path}")
    else:
        print("Excel報表生成失敗") 